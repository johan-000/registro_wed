from flask import Flask, request
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
import os, base64
from flask import render_template_string
from openpyxl.styles import Border, Side


app = Flask(__name__)

CARPETA_FIRMAS = r"C:\\Users\\johan\\OneDrive\\Escritorio\\firmas"
ARCHIVO_EXCEL = r"C:\\Users\\johan\\Downloads\\GIC-F-LMPM-048 Control de Ingreso de Visitantes y Personal de Laboratorio.xlsx"
os.makedirs(CARPETA_FIRMAS, exist_ok=True)


# Logo en base64 (verde SENA)
logo_base64 = (
    "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAACdUlEQVR4nO3ZMQ6DMBBA0Vn6/39KrqWUR2TZ8WTpjAVSu9G3MiAoSAAAAAAAAAAAwBPkDaEILXKBoV1pULXTZkxw7mFUz8Fss4W1KMdAxyNWDgIbsq+w5PgNP2BlPXG1XIbYgaZXqDlfUm0iPgFV4Df1aZxD6AqzljTqskV8G3Rj7T7EX8B0uZQX+GljlZx/0uwjl+S7HgU04CMt3MbUAVaRxD0wIkYt1BdAvKoMW8PKf5uk2XwBNzEMjWvGUS5HkaT8BPG7Hz/NH8AiPwqMkG6GflKk3i8r3UazewAYZ8UMXcrx+KflnDj0DeovQZyfuE8FPGLZj/YGv5+4TS9rAf2Tsk5B3UUl7TxR8a87eT6eHTeec9Gx5+rncNgbLRblmVpdtm4bepbNmF2AD/qKqAIOtfdUPYwA/h2fHwAAAAAAAAAAADAf/gDZjoWvDFcgnQAAAABJRU5ErkJggg=="
)



def escribir_en_excel(data, firma_path):
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb.active
    fila = 8
    while ws.cell(row=fila, column=5).value not in (None, ""):
        fila += 1

    fecha = datetime.now()
    meses = ["", "Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    ws.cell(row=fila, column=1, value=fecha.year)
    ws.cell(row=fila, column=2, value=meses[fecha.month])
    ws.cell(row=fila, column=3, value=fecha.day)
    ws.cell(row=fila, column=4, value=fecha.strftime('%H:%M'))  # Hora de ingreso

    ws.cell(row=fila, column=5, value=data['nombres'])
    ws.cell(row=fila, column=6, value=data['documento'])
    ws.cell(row=fila, column=7, value=data['motivo'])
    ws.cell(row=fila, column=8, value=data['eps'])
    ws.cell(row=fila, column=9, value=data['arl'])
    ws.cell(row=fila, column=10, value=data['autoriza'])
    ws.cell(row=fila, column=11, value=data['acudiente'])
    ws.cell(row=fila, column=12, value=data['celular'])

    # Firma en columna 13 (M)
    img = XLImage(firma_path)
    img.width = 120
    img.height = 50
    ws.add_image(img, f"{get_column_letter(13)}{fila}")
    ws.row_dimensions[fila].height = 60

    # Columna 14: hora de salida
    ws.cell(row=fila, column=14, value="")  # Puedes dejarlo vacío para llenar después

    # Columna 15: observaciones (si ya vienen del formulario)
    ws.cell(row=fila, column=15, value=data.get("observaciones", ""))

    # Aplicar bordes a columnas 1 a 15
    borde = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for col in range(1, 16):  # A (1) hasta O (15)
        ws.cell(row=fila, column=col).border = borde

    wb.save(ARCHIVO_EXCEL)
@app.route("/")
def inicio():
    return render_template_string("""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {
                font-family: Arial, sans-serif;
                background: #f4f4f4;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            }
            .container {
                background: white;
                padding: 30px;
                border-radius: 12px;
                text-align: center;
                box-shadow: 0 0 15px rgba(0, 0, 0, 0.1);
            }
            .logo {
                width: 100px;
                margin-bottom: 20px;
            }
            h1 {
                color: #007C3E;
                font-weight: bold;
            }
            .btn {
                background: #007C3E;
                color: white;
                padding: 10px 20px;
                margin: 10px;
                border: none;
                border-radius: 8px;
                text-decoration: none;
                display: inline-block;
                font-size: 16px;
            }
            .btn:hover {
                background: #005f2d;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <img src="data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMwAAADACAMAAAB/Pny7AAAAflBMVEX///8ArwAArgDL7cv1/fWE04Q+vT4AswD7/vtCwkIArADH7Mfd9N35/vkAsQDx+/G46Lic3pwctByn5KeD2IMquSqO3I6z5rNezV4mvSbr+evZ89ls0Gzk9+R60HrT8dPB68GT25Or46tNv01403hkzWQ2vDZNxk1VyFVuzG5SjlzMAAAP60lEQVR4nO1diZKqSBCUEgdQDkURRRFwBPX/f3DxbiqbwwMdN15GbMS+EaWT7q67i06nZeg3tH2rluEs/Gk6+z0i3Yy7gfXpET2MYLDKbI36dESftPlubH4nHTXOqeQklCuIbHc2Nj89sAdgJgcqCgMp3jL+uslZpC5SOdExxuqnR3cf4q0tpXKanGnw6fHdAT3eyqflTMedOp8eYnNE2woqx8lJvmalmalSTUYhb/0lUsAZudVUctiz6NPDbIaB0a8lQ8p3bBt1UyrIRDbG4Btsta7XgItC9jdMjTVpMjE5si/YNYHRjAvR+tNDrcfQbUZG6W/+/DrTJ1pDMmT8+XWm7xpyyUVA/OnB1kH16pXMmczf3zTOvOnMKDT+9GDr4DTSMv/IvB//LzJ37JnRpwdbBydrLs3+/Myos8Z6Rht+erC1aGpnfoMF0Nw2o83fj9GEDfzMIxdl/Pe9M2vazNKkb/BnOlEj4Ux27897ADmcTZOp+ZIYgD5o4GsS9b4jDKiOGgi03eLTw2yIcF8b0cyG37DIDtAHsxou7ug7FtkBdVkAu/f39eUNerwrZUPkJt/EJWezSLWSzBnNvy1z1ukEidaX5DRJ232FgmGwukuNinSIbGP8XUvsCrU7yXKVQ2corrtbh58e1OOwzOE03f4c8bsZR186K1c44SI6wgy/JPP3D//wD/8z6EHgBOoXamsBerCIuvEoGfXS2Wy2T5JkPOguQqfAyhnU4qgd9bDL/myyp6OH4qcyPeQUf2LR+PGqZne9NzxbO+CgtU//5xrpdBgJAQhfq4M9PdxUH7ns7ymLyThT8VNJ+knt2YXfNQYNqUSj2dymq/lxtKZOsDXFSAZXa9e/flCG/uRMhn+wL1YCOj1b+BBD6bqvsR9eNorqBKOtplCJH3L4+8/osgyG9QGYKxnumO0L5lk+M7fPNAylRzw6kl/UwJIwJ24hTHydndsfrnVij5PJ2RTCsjVknCWErqnBQmM570Ox66nmVfyj5zckQ1RGJh/yVHi01WT0MYbh86dRa3rH4vjI9maTaa/XW6Ve4efSQEamcs8gUVsIZ1STkYbgyB3VSbRNX7x8cwlr6dFUCLdSNgQy5C6TKeK4FqRkFJrfRlNJJlwKXxL+t3ahpSIZzxe4J7bwwSl0XyCTlf+0nMzh9y8rrYqMmly/TfPd7ZnWL7QCGdsYdZ1L8X44WV0xPamJJ8ko/WsQsIKMPsxu498sVrcBkl1TFbkspFVyOr31sGsGVk4p/8+6oCMhE1sAHcmQLUrL2TmsUUFmcSv1PDywrvBc6Lc6H5IofE/b7jZdjdZ+14TQUHHPbEaAtYNk5kvjyoZoF9WQcZLb1W5iHUyB228pq0oP3ORFLwctY9u2623T3sgvmk0FaWYrtsLhmUjG831h3dhLs5KMLlzcP14b7cQnuK6UaGMX83cnU0ZRXM/YJ/FtgoqiWSKZZWTmQ3WkiWwWVWQWW+ExHHeYtRaqcumncqFZ43m/L92up1lSss01dF+nNCmTkumEQjFqzkYtJyNUR5J9XlOhaA7QplIG6N1VZkuikDdG2sZ8jkwn3IsLf+OUkhnfxkGXXI4+FGpZSKmpiVAHo72Rr7ZD4E46xvNCr7MAlL58mXUKIupYQavKyUSC6PKu20NQPIfFV5uvUqN1st9lrqucQpGMjXZKqxbJuB6Huw1LyOTbWJgbbeQkMjJOKtw57QbhGf6POJh6G+3wU9FwnKzSrZdLMzZH9BPrjAx5PZ9jfRYWEjKFlCf9jFfiozqTsUainsjSg897QDoTa6bILasksMxocYEZ6B3dWQxySptt0dAkJXE4mYYWwJlM7tll4pPIbiLqSiY2ylaxUvj7b1d+2+7yd3bF6OLMWWHX34gWCdmTp8noY/Hbgo66kBHtyyoQlahOf963L1BSwbXVndGLyXTUqdxiO5MJSj6WfKHk5MpQMEqVrFA8tS6Q6T1PphPIi59PZET7shY7qeos1O9QNonPkkId9Aqbzjt64EXbLJ3IEJeTOSwkmZ9zJGMKkuxyxpNJfuEb8qKVSWGfa3PP2O6WWyObiwKAlOWRJLPNZHekiVVOpqjMC2SClWDyaNsZQqzNobnURlv8MHOrnxs3h/8Ki2wbd4CMFKT0qsh0ohnOzYGMPha1fBqFCLE8j8iQqs7q3P3xi256loUNojM1ZPLb4Q1yMmZBbseygQom6GEhSs9H6IOlVzy+K/I4ujfJ5Sk8PzMdK9/nYF+snY1o7cg3hFVwX3PVKbuoY673PwdvEHecZrvZct29ykG/ngxxMn1WYFpYUBcyoo+gbEusfHMpPAbql9hoVhitJzvDLUZ3Nc9YTv2umJIcbI0aZMZB7On+7cJsx0S4M+Lf2fqz7PbP7bgkEKsPd+K3Mr/MT9Od0Iy68Xp8gd892DfFJEBu/NTjOBJH/AtXcCp+R/xHWBpUVsPCt+py2P+f/gr/8A//UAGz2yK4AHOiFm8WBZ3VT3sATRGlbd7N76RSK/sl6ENE2Elqk6GP380edZbNPaE7IauRjWrqap+63bhFMgXv+4J1U5f4brRKhrJYIm+CTUu3a5XMMfgqwT0e/n03bJNMyfkYtVdXW//oDdsjQ1pSYqIutk3PC955xxbJbEsDwaMGPvcjd2yNDNnlR8rDXStT0xoZ0nYVqaBh09OP992zNTJuXM6lY6XfRKbuENagDfF8JNPCAqZtdVJLTzA//fxNczItTHlulNWUhEVlnTZq4pA1d+2sew9j75XcmxtlasB0jjUuS2r8Pj6cabdjqQ8j9Jey5UIZy5pYfo8rHVOeT/JG0ePDsZ4MJJmynWzzfNZi5/I4nTQJQ3aZ1fAeSA4yg7VsTW0l5RLBmeCk0vazbSmhHlTS8K87p9xS4yIBi/wkF70XAZgmkDBxDsKffmL2TSvh/elIVtb8TsQQyQejbH1YiUQz7t2EXDyTtvrsMUHItNKMCa7wlBqSWJ5Y85x99MDzgp/9Jy0uXqFOz75YH/J1AbekyJ68a+ASOBM+nP6e7eHBJVVK2oTL3RgexfyD3RuGsGP4aMLNLXMGGUqdt0IsSVC+BUGqMC6nsxkCfDF3DEcSIiildP23jb4IyamFHdN6i7QwUqiOT7ibRr8fOpW6ANlqM6vFSYp0IeNqzmBqmpzBeD3E0tzTQLQlM8q63ADjikT3YWqMkrKrdhH/8HGcCutuCIAuJIYDkAHuJ1q4BCulCOhZqvvY62zJNpVe/0jeAMlIt2yBhFgegz6oM+VT0+B4zKthprA+WAJDfnbmt8ueO3jQ5yqwN0Ifa2yY9oyZK2YGVGT1MCp2d969eWrwefLifGsiDZDTnKtFk4dUyK6Lh7wYExgkF8tgeF0unLELdcg/vbmN2wAs/3lcvEItbdjU54cxVR7cIFq+cWqclPuXfRaK0NelWRg0OAdYq8UvaRFrnBimPyqa6ROl7LlbaKJt3+ZzLrjjT1pSvELtlVFRZL5+ICmjexMXia3LQpi5Xq8gcztDcsWIL8q6aPXLMAALxGUNvsJ9ZWCc3ClbaAGK5/c0cwvAV7Y39UZZ8RtGzH50CB70z1usZ9/jQ8vYfaNdJZUD/T1TNg72En1Hm01oK0c2e8uHk9TX1nrcF4ig/abXvgcNyQgibpTBnpKQAacTGwn2ZdUqr0XE35IDRlmwbJAaz6eTGZwmRNFcfsmrAUuIuL+l+41KfeiHSUBrDcqGO0ivRsy9eui7KMnyH1okgJtmp2yD55uRXzJpVQYEqEBWbC1wzzGn4s32M8gYwslxiQedtepB+6AOeJ8BSfR5tzYDc7zljwEMfbCBUIK/EiEcE7GZUabv0WY89s3R0fSEMHkEUzNvsQ36mK+VPvdwoZKE5pepg67UYOjrUCJEWWtTE/Fm7KQwxYYJJOX6+HVuBuWPgg014LNH/WlLXKwNiKSUBWQShV8hVGpBapoU3qYk5hPbnzfsZnQvMIHBfUbYF5cGIidAExnasqHqkFegZSvWcwglQ8Syw8EEtkWhl1EARwBtLtghOp27Fy1w0VFDM9mqD3mkjCkKHeeWB59VcPyUtIUoGpRXQ/YBXDKwrhysZuLWpGlgpPTlmhOtWu4rW3ggDtxj3FQ8+GxJMlgv96C7XF8Sr4eRBDnhvH6hp8cJ3IQI0YOevlgGBCvILDErUAVfkXUyOw11A64d8yixROjlSY4YfJQf9kThxXPy8nkUEnOeO4Qkx/lA9auAc+8yX1myOiay1eHww/5k8wpoWNEKVHg9A2sMcS2+KxPYtz/yfYtpXf6mUKddD3oBy5ibswFk9CX9/U6AikYIupugBV5YVWdNYWK4JoN1TtyRvMIBZ1+ZsQW5RrfpZUkOeNUHO0B+iOPDxJQbiJKyLm58Y4nQ6kVcVDTKlsUrHF54Stq+fF3oPUiI8IqnAbpF8WvIQKqlz8t9cBsYVTs2wGJTrhfR94Hiu4cQggliM4/J5C+dIq26+xh6lNyZMDGU8AoP2sFaSuYrS1QHT3AyBODmgW08goVWIurvwgAmZs4SQQODs619hwnKAJ78D/EM5PMmmsPzyhD/yc02doVb+94PB4S9wvou6ZDkUCAPcjdqa1ssHzTqrj6sClkPqLwJIGgFbum9CPEnmcNlgtte2qlLgMT5MYoPSccEXXmzjEbAHBjPaGGsC0vlZQj3YAcwm0FSIsRrou7DAlYDL/eBQ0olnW8A6AtozOmMeET3uTLB+vCCzoUsaZtmQgf7gpHHAiSvTXJ0oY21x7ShJJLe1CRcgMzXWKLXBCvpiTe9qZgDY2JZYh00j6WgEaQxgxM9aKn32gQ6dGTiv6WjnuZHASoQgAwgg6sw9ixrbYtSWClreEoKK0bo8hTxfRU8A1ijCpMBKF4eLRE6dIYq/BY/RWFByMZuuPvPN8AsG2vJpyb8Bg870OrQKLZqY5klTHDe6RCaIGDsvVNxRS4A+BsfmkOPxR5kPJ0iyajeW48o8QWYDytWgxEczbsPg9vckM36k/EgCml3V4qqK0hNs7BPcDvh28CArYYe7y72rc2OxqCteP/bC3X0Bfj5u2sFJXnTp3OC0bmNL98PDtTGP5K2B8cOjtpdjhzkHs8LQs6LYzl2vvsLE4NZe+6RNEOEkVBWcX7KHPS919QHn9iwch8sbnrstZISX4A5nero0KjzZWXoOZs+6/Kq+7Vpo4bAqkGbHaTNPabcJnxZWiNcacyVMLGC/8HyMH2Iyqbo7ecP7mf4whSNmQyLE4O6++HKEBWr11lpIb/7i4EV597j+XowvXNl88ajtA70qug/865vrA94ScSvIaDiXDGe0WeSZ+O97TC9CRXncI7kPkSwauku8/sJQOk+QQXdvYDkLhicbUES0Hq2mNJEGXCHy/oEQgjI2E+/UlpyIIhngNvBwGD+pz17frfyGiFSjLe8EDkY57aaeGeoHn8EBW+fKJvyU4NN8B+QO0Ruriu6iAAAAABJRU5ErkJggg==" class="logo">
            <h1>Registro de Asistencia </h1>
            <a href="/registrar" class="btn">Registrar Ingreso</a>
            <a href="/salida" class="btn">Registrar Salida</a>
            
        </div>
    </body>
    </html>
    """, logo=logo_base64)

@app.route("/registrar", methods=["GET", "POST"])
def registrar():
    if request.method == "POST":
        datos = {
            'nombres': request.form["nombres"],
            'documento': request.form["documento"],
            'motivo': request.form["motivo"],
            'eps': request.form["eps"],
            'arl': request.form["arl"],
            'autoriza': request.form["autoriza"],
            'acudiente': request.form["acudiente"],
            'celular': request.form["celular"],

                }

        firma_data = request.form["firma"]
        nombre_firma = f"{datos['documento']}_{datetime.now().strftime('%H%M%S')}.png"
        ruta_firma = os.path.join(CARPETA_FIRMAS, nombre_firma)

        with open(ruta_firma, "wb") as f:
            f.write(base64.b64decode(firma_data.split(",")[1]))

        escribir_en_excel(datos, ruta_firma)
        return mostrar_mensaje_exito("✔ Registro de ingreso exitoso")

    return render_template_string("""
<!DOCTYPE html>
<html lang="es">
<head>
    <meta charset="UTF-8">
    <title>Registrar Ingreso - SENA</title>
    <style>
        body { font-family: Arial, sans-serif; background-color: #f4f4f4; margin: 0; padding: 0; }
        header { background-color: #007C3E; color: white; padding: 20px; text-align: center; }
        .container { max-width: 700px; margin: 30px auto; background: white; padding: 30px;
            box-shadow: 0 0 15px rgba(0,0,0,0.1); border-radius: 10px; }
        input[type="text"], input[type="tel"], select {
            width: 100%; padding: 10px; margin: 10px 0;
            border: 1px solid #ccc; border-radius: 5px; box-sizing: border-box;
        }
        button {
            background-color: #007C3E; color: white; padding: 10px 20px;
            border: none; border-radius: 5px; cursor: pointer; margin-top: 10px;
        }
        button:hover { background-color: #005f2d; }
        canvas { border: 1px solid #ccc; margin-top: 10px; }
        a { display: inline-block; margin-top: 20px; color: #007C3E; text-decoration: none; }
        .modal {
            display: none; position: fixed; z-index: 999; left: 0; top: 0;
            width: 100%; height: 100%; background-color: rgba(0,0,0,0.5);
            justify-content: center; align-items: center;
        }
        .modal-content {
            background-color: #fff; border-radius: 10px; padding: 20px;
            width: 90%; max-width: 500px; box-shadow: 0 0 10px #333; color: #333;
        }
        .modal-header {
            font-weight: bold; font-size: 20px; margin-bottom: 15px; color: #007C3E;
        }
        .modal-buttons {
            text-align: right; margin-top: 20px;
        }
        .modal-buttons button { margin-left: 10px; }
        .alerta {
            background-color: #f44336; color: white; padding: 10px;
            margin: 10px 0; border-radius: 5px; display: none;
        }
    </style>
</head>
<body>
<header>
    <h1>Registro de Ingreso en el Laboratorio - SENA</h1>
</header>
<div class="container">
    <div class="alerta" id="alertaError"></div>
    <form method="POST" id="formulario">
        <label>Nombre y Apellido:</label>
        <input type="text" name="nombres" id="nombres">

        <label>Documento de Identificación:</label>
        <input type="text" name="documento" id="documento">

        <label>Motivo de Visita:</label>
        <select name="motivo" id="motivo">
            <option value="" disabled selected>Seleccione un motivo</option>
            <option value="Transferencia de conocimiento">Transferencia de conocimiento</option>
            <option value="Asesoría de gestión de mantenimiento">Asesoría de gestión de mantenimiento</option>
            <option value="Solicitud de servicios del laboratorio">Solicitud de servicios del laboratorio</option>
        </select>

        <label>EPS:</label>
        <input type="text" name="eps" id="eps">

       <label>ARL:</label>
        <select name="arl" id="arl">
      <option value="" disabled selected>Seleccione una ARL</option>
      <option value="SURA">SURA</option>
      <option value="Colpatria">Colpatria</option>
     <option value="Positiva">Positiva</option>
      <option value="AXA Colpatria">AXA Colpatria</option>
     <option value="Bolívar">Bolívar</option>
     <option value="La Equidad">La Equidad</option>
     <option value="Alfa">Alfa</option>
     <option value="Ninguna">Ninguna</option>
     </select>

        <label>Quién autoriza el ingreso:</label>
        <select name="autoriza" id="autoriza">
            <option value="" disabled selected>Seleccione quien autoriza</option>
            <option value="Personal técnico">Personal técnico</option>
            <option value="Responsable del laboratorio">Responsable del laboratorio</option>
            <option value="Responsable del sistema de gestión">Responsable del sistema de gestión</option>
        </select>

        <label>Persona en caso de emergencia:</label>
        <input type="text" name="acudiente" id="acudiente">

        <label>Teléfono de persona en caso de emergencia:</label>
        <input type="tel" name="celular" id="celular">

     
        <label>Firma:</label><br>
        <canvas id="canvas" width="300" height="100"></canvas><br>
        <button type="button" onclick="limpiarCanvas()">Limpiar Firma</button><br>

        <input type="hidden" name="firma" id="firma">
        <button type="button" onclick="validarYMostrar()">Registrar</button>
    </form>
    <a href="/">Volver al Inicio</a>
</div>

<div class="modal" id="modal">
    <div class="modal-content">
        <div class="modal-header">¿Los datos ingresados son correctos?</div>
        <div id="resumen"></div>
        <div class="modal-buttons">
            <button onclick="cerrarModal()">No, corregir</button>
            <button onclick="enviarFormulario()">Sí, registrar</button>
        </div>
    </div>
</div>

<script>
function mostrarError(mensaje) {
    const alerta = document.getElementById("alertaError");
    alerta.innerText = mensaje;
    alerta.style.display = "block";
    setTimeout(() => alerta.style.display = "none", 4000);
}

function validarYMostrar() {
    const campos = [
        { id: "nombres", nombre: "Nombres" },
        { id: "documento", nombre: "Documento de Identificación" },
        { id: "motivo", nombre: "Motivo de Visita" },
        { id: "eps", nombre: "EPS" },
        { id: "arl", nombre: "ARL" },
        { id: "autoriza", nombre: "Quién autoriza el ingreso" },
        { id: "acudiente", nombre: "Persona en caso de emergencia" },
        { id: "celular", nombre: "Teléfono de persona en caso de emergencia" }
    ];

    for (let campo of campos) {
        const valor = document.getElementById(campo.id).value.trim();
        if (valor === "") {
            mostrarError("El campo '" + campo.nombre + "' es obligatorio.");
            return;
        }
    }

    const canvas = document.getElementById("canvas");
    const ctx = canvas.getContext("2d");
    const pixel = ctx.getImageData(0, 0, canvas.width, canvas.height).data;
    let vacio = true;
    for (let i = 0; i < pixel.length; i++) {
        if (pixel[i] !== 0) {
            vacio = false;
            break;
        }
    }
    if (vacio) {
        mostrarError("Por favor dibuje su firma antes de continuar.");
        return;
    }

    mostrarConfirmacion();
}

function guardarFirma() {
    var canvas = document.getElementById("canvas");
    var dataUrl = canvas.toDataURL();
    document.getElementById("firma").value = dataUrl;
}

function limpiarCanvas() {
    var canvas = document.getElementById("canvas");
    var ctx = canvas.getContext("2d");
    ctx.clearRect(0, 0, canvas.width, canvas.height);
}

function mostrarConfirmacion() {
    guardarFirma();

    const campos = [
        { id: "nombres", nombre: "Nombres" },
        { id: "documento", nombre: "Documento de Identificación" },
        { id: "motivo", nombre: "Motivo de Visita" },
        { id: "eps", nombre: "EPS" },
        { id: "arl", nombre: "ARL" },
        { id: "autoriza", nombre: "Quién autoriza el ingreso" },
        { id: "acudiente", nombre: "Persona en caso de emergencia" },
        { id: "celular", nombre: "Teléfono de persona en caso de emergencia" },

    ];

    let resumen = "<ul style='list-style:none; padding-left:0'>";
    campos.forEach(campo => {
        const valor = document.getElementById(campo.id).value.trim();
        resumen += `<li><strong>${campo.nombre}:</strong> ${valor}</li>`;
    });
    resumen += "</ul>";

    document.getElementById("resumen").innerHTML = resumen;
    document.getElementById("modal").style.display = "flex";
}

function cerrarModal() {
    document.getElementById("modal").style.display = "none";
}

function enviarFormulario() {
    document.getElementById("modal").style.display = "none";
    document.getElementById("formulario").submit();
}

window.onload = function () {
    const canvas = document.getElementById("canvas");
    const ctx = canvas.getContext("2d");
    let dibujando = false;

    canvas.style.touchAction = "none"; // Evita scroll en móvil

    function getPos(e) {
        const rect = canvas.getBoundingClientRect();
        if (e.touches && e.touches.length > 0) {
            return {
                x: e.touches[0].clientX - rect.left,
                y: e.touches[0].clientY - rect.top
            };
        } else {
            return {
                x: e.offsetX,
                y: e.offsetY
            };
        }
    }

    function empezarDibujo(e) {
        dibujando = true;
        const pos = getPos(e);
        ctx.beginPath();
        ctx.moveTo(pos.x, pos.y);
    }

    function dibujar(e) {
        if (!dibujando) return;
        e.preventDefault();
        const pos = getPos(e);
        ctx.lineWidth = 2;
        ctx.lineCap = "round";
        ctx.lineTo(pos.x, pos.y);
        ctx.stroke();
        ctx.beginPath();
        ctx.moveTo(pos.x, pos.y);
    }

    function terminarDibujo() {
        dibujando = false;
        ctx.beginPath();
    }

    // Mouse events
    canvas.addEventListener("mousedown", empezarDibujo);
    canvas.addEventListener("mousemove", dibujar);
    canvas.addEventListener("mouseup", terminarDibujo);
    canvas.addEventListener("mouseleave", terminarDibujo);

    // Touch events
    canvas.addEventListener("touchstart", empezarDibujo);
    canvas.addEventListener("touchmove", dibujar);
    canvas.addEventListener("touchend", terminarDibujo);
};

</script>
</body>
</html>
""")


def mostrar_mensaje_exito(mensaje):
    return render_template_string("""
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Registro Exitoso</title>
        <style>
            body {
                background-color: #e8f5e9;
                font-family: Arial, sans-serif;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            }
            .contenedor {
                background: white;
                padding: 30px;
                border-radius: 15px;
                box-shadow: 0 5px 20px rgba(0, 128, 0, 0.2);
                text-align: center;
            }
            .icono {
                font-size: 60px;
                color: green;
            }
            h2 {
                color: #2e7d32;
            }
            a {
                margin-top: 20px;
                display: inline-block;
                padding: 10px 20px;
                background-color: #2e7d32;
                color: white;
                text-decoration: none;
                border-radius: 5px;
            }
            a:hover {
                background-color: #1b5e20;
            }
        </style>
    </head>
    <body>
        <div class="contenedor">
            <div class="icono">✔</div>
            <h2>{{ mensaje }}</h2>
            <p>Gracias por registrar su salida.</p>
            <a href="/">Volver al inicio</a>
        </div>
    </body>
    </html>
    """, mensaje=mensaje)


def mostrar_mensaje_error(mensaje):
    return render_template_string("""
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Documento No Encontrado</title>
        <style>
            body {
                background-color: #e8f5e9;
                font-family: Arial, sans-serif;
                display: flex;
                justify-content: center;
                align-items: center;
                height: 100vh;
                margin: 0;
            }
            .contenedor {
                background: white;
                padding: 30px;
                border-radius: 15px;
                box-shadow: 0 5px 20px rgba(0, 128, 0, 0.2);
                text-align: center;
            }
            .icono {
                font-size: 60px;
                color: red;
            }
            h2 {
                color: #2e7d32;
            }
            a {
                margin-top: 20px;
                display: inline-block;
                padding: 10px 20px;
                background-color: #2e7d32;
                color: white;
                text-decoration: none;
                border-radius: 5px;
            }
            a:hover {
                background-color: #1b5e20;
            }
        </style>
    </head>
    <body>
        <div class="contenedor">
            <div class="icono">✘</div>
            <h2>{{ mensaje }}</h2>
            <p>Por favor verifique el número de documento.</p>
            <a href="/salida">Intentar de nuevo</a>
        </div>
    </body>
    </html>
    """, mensaje=mensaje)

@app.route("/salida", methods=["GET", "POST"])
def salida():
    if request.method == "POST":
        documento = request.form["documento"]
        observaciones = request.form.get("observaciones", "")

        wb = load_workbook(ARCHIVO_EXCEL)
        ws = wb.active

        fila = 8
        encontrada = False
        while ws.cell(row=fila, column=6).value:
            if str(ws.cell(row=fila, column=6).value) == documento and ws.cell(row=fila, column=14).value in (None, ""):
                encontrada = True
                hora_salida = datetime.now().strftime("%H:%M")
                ws.cell(row=fila, column=14, value=hora_salida)
                ws.cell(row=fila, column=15, value=observaciones)

                # ✅ Aplicar bordes a toda la fila (1 a 15)
                from openpyxl.styles import Border, Side
                borde = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                for col in range(1, 16):
                    ws.cell(row=fila, column=col).border = borde

                break
            fila += 1

        if encontrada:
            wb.save(ARCHIVO_EXCEL)
            return mostrar_mensaje_exito("✔ Registro de salida exitoso")
        else:
            return mostrar_mensaje_error("❌ No se encontró una salida pendiente para esa cédula")

    # Interfaz HTML
    return render_template_string("""
    <!DOCTYPE html>
    <html lang="es">
    <head>
        <meta charset="UTF-8">
        <title>Registro de Salida</title>
        <style>
            body {
                font-family: Arial, sans-serif;
                background-color: #f4f4f4;
                padding: 30px;
            }
            h2 {
                color: #007C3E;
                text-align: center;
            }
            form {
                background: white;
                padding: 20px;
                border-radius: 8px;
                max-width: 400px;
                margin: auto;
                box-shadow: 0 0 10px rgba(0,0,0,0.1);
            }
            input {
                width: 100%;
                padding: 8px;
                margin: 10px 0;
                border-radius: 4px;
                border: 1px solid #ccc;
            }
            button {
                background-color: #007C3E;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 4px;
                cursor: pointer;
            }
            button:hover {
                background-color: #005f2d;
            }
            a {
                display: block;
                margin-top: 20px;
                text-align: center;
                color: #007C3E;
            }
        </style>
    </head>
    <body>
        <h2>Registrar Hora de Salida</h2>
        <form method="POST">
            <label>Documento de Identidad:</label>
            <input name="documento" required>
            <label>Observaciones:</label>
            <input type="text" name="observaciones" id="observaciones">
            <button type="submit">Registrar Salida</button>
        </form>
        <a href="/">Volver</a>
    </body>
    </html>
    """)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
#al momento del mensaje de registro exitoso quiero una interfaz grafica mas bonita verde dirigida al sena con logo y lo mismo para el registro extitoso ded salida